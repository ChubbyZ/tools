import subprocess
import re
import openpyxl

# 代理服务器的IP地址和端口号
#proxy_ip = 'YOUR_PROXY_IP'
#proxy_port = 'YOUR_PROXY_PORT'

# 从txt文件中读取域名列表
with open('domain_list.txt', 'r') as file:
    domain_list = file.read().splitlines()

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 添加表头
sheet['A1'] = 'Domain'
sheet['B1'] = 'IP'

# 设置代理环境变量
#proxy_env = {'http_proxy': f'http://{proxy_ip}:{proxy_port}', 'https_proxy': f'http://{proxy_ip}:{proxy_port}'}

row = 2  # 从第2行开始写入结果

for domain in domain_list:
    try:
        # 执行nslookup命令
        command = f'nslookup {domain}'
        result = subprocess.run(command, shell=True, capture_output=True, text=True)

        # 提取IP地址
#        ip_match = re.search(r'Address:\s+([\d.]+)', result.stdout)
#        if ip_match:
#            ip = ip_match.group(1)
#            sheet.cell(row=row, column=1).value = domain
#            sheet.cell(row=row, column=2).value = ip
#        else:
#	   print
        non_auth_match = re.search(r'Non-authoritative answer:\nName:\s+\S+\nAddress:\s+([\d.]+)', result.stdout)
        if non_auth_match:
            ip = non_auth_match.group(1)
            sheet.cell(row=row, column=1).value = domain
            sheet.cell(row=row, column=2).value = ip
        else:
            sheet.cell(row=row, column=1).value = domain
            sheet.cell(row=row, column=2).value = '无法解析'
    except subprocess.CalledProcessError:
        sheet.cell(row=row, column=1).value = domain
        sheet.cell(row=row, column=2).value = '无法解析'

    row += 1

# 保存Excel文件
workbook.save('nslookup_results.xlsx')

print('操作完成！结果已保存到nslookup_results.xlsx文件中。')